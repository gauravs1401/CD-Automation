import numpy as np
import pandas as pd
import win32com.client
import json
import subprocess
import time
from datetime import datetime
import psutil
import openpyxl
from openpyxl.styles import PatternFill
import shutil
import os
import re

var_2 = (datetime.now()).strftime("%d.%m.%Y")

with open('Cred\\Details.json') as f:
    contents = json.load(f)

path = r'C:\\Program Files (x86)\\SAP\\FrontEnd\\SapGui\\saplogon.exe'
subprocess.Popen(path)
time.sleep(1)

# Connect to SAP GUI
SapGuiAuto = win32com.client.GetObject("SAPGUI")
if not type(SapGuiAuto) == win32com.client.CDispatch:
    raise Exception("SAP GUI not found")

application = SapGuiAuto.GetScriptingEngine
if not type(application) == win32com.client.CDispatch:
    raise Exception("SAP GUI scripting engine not found")

# Connect to SAP system
connection = application.OpenConnection(contents['DCPConnectionName'], True)
if not type(connection) == win32com.client.CDispatch:
    raise Exception("SAP system not found")

session = connection.Children(0)
if not type(session) == win32com.client.CDispatch:
    raise Exception("SAP session not found")

# Set up event handling
if hasattr(session, "On"):
    session.On("ScriptingCommandBarPanel", session.Events)

# Login to SAP system
session.findById("wnd[0]/usr/txtRSYST-BNAME").text = contents['user']
session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = contents['password']
session.findById("wnd[0]").sendVKey(0)

# Navigate to sales order transaction
session.findById("wnd[0]/tbar[0]/okcd").text = "/nVA02"
session.findById("wnd[0]/tbar[0]/btn[0]").press()


def cd_remove(case_number, order):
    # Entering sales order number
    session.findById("wnd[0]/usr/ctxtVBAK-VBELN").text = order
    session.findById("wnd[0]/usr/ctxtVBAK-VBELN").setFocus()
    session.findById("wnd[0]/usr/ctxtVBAK-VBELN").caretPosition = 10
    session.findById("wnd[0]").sendVKey(0)

    # Capturing the "Product Table" in the main screen and updating "flag_2" to check if we have an "ALLETRA" or "3PAR" product
    Product_Table = session.findById(
        "wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/subSUBSCREEN_TC:SAPMV45A:4900/tblSAPMV45ATCTRL_U_ERF_AUFTRAG")

    # Setting the number of visible rows to display at once
    Visible_Rows = Product_Table.VisibleRowCount

    list_Product_Descriptions = []

    flag_1 = False
    flag_2 = 0
    i = 0
    while i < (round(Product_Table.RowCount / Visible_Rows)) and (flag_1 == False):  # length of the order

        i = i + 1
        for row_index in range(Visible_Rows):

            try:
                Product_Table = session.findById(
                    "wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/subSUBSCREEN_TC:SAPMV45A:4900/tblSAPMV45ATCTRL_U_ERF_AUFTRAG")
                cell = Product_Table.GetCell(row_index, 5)
                if cell is not None:
                    product_name = cell.Text
                    if 'Alletra' in product_name or '3PAR' in product_name:
                        print(f"{product_name}")
                        flag_1 = True
                        flag_2 += 1
                        break
                    list_Product_Descriptions.append(cell.text)
                    # print(f"Product name at row {row_index + 1}: {product_name}")

            except:
                # print(f"No valid product name found at row {row_index + 1}")
                pass

        session.findById("wnd[0]").sendVKey(82)
        Product_Table = session.findById(
            "wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/subSUBSCREEN_TC:SAPMV45A:4900/tblSAPMV45ATCTRL_U_ERF_AUFTRAG")

    # Creating "Fallout" & "Orders" DataFrames if there are no orders in either of them
    df_Fallout = pd.DataFrame(columns=['CASE NUMBER', 'HPON', 'COMMENTS'])
    df_Orders = pd.DataFrame(columns=['CASE NUMBER', 'HPON', 'COMMENTS'])

    # Going to the second screen to check if the order is "ARUBA" or "HYBRID-IT"
    session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/btnBT_HEAD").press()
    text_Aruba = session.findById(
        "wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4301/txtTVKBT-BEZEI").text

    session.findById("wnd[0]/tbar[0]/btn[3]").press()

    # First check, if the order is ARUBA or HYBRID IT
    if text_Aruba == 'Aruba':
        session.findById("wnd[0]/tbar[0]/btn[3]").press()  # Going back to variant "VA02"
        print("ARUBA ORDER")

        # Dictionary for Fallout Order
        dict_Fallout = {'CASE NUMBER': [case_number], 'HPON': [order], 'COMMENTS': 'FALLOUT/ARUBA'}
        df_Fallout = pd.DataFrame(dict_Fallout)

    # Second check, if the order has "ALLETRA" or "3PAR" products
    elif flag_2 != 0:

        session.findById("wnd[0]/tbar[0]/btn[3]").press()  # Going back to variant "VA02"
        # Dictionary for Fallout Order
        dict_Fallout = {'CASE NUMBER': [case_number], 'HPON': [order], 'COMMENTS': 'FALLOUT/ALLETRA|3PAR'}
        df_Fallout = pd.DataFrame(dict_Fallout)

    # Third check, "HEADER STATUS"
    else:

        # Checking the "HEADER STATUS"
        session.findById("wnd[0]/tbar[1]/btn[22]").press()  # Going to "HOLD CODE CHANGE LOG"
        text_header_status = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell").GetCellValue(0,
                                                                                                                      "HSTATUS")

        session.findById("wnd[0]/tbar[0]/btn[3]").press()  # Going back to the first screen

        # if text_header_status in list_Exclusion_Status:
        if any(word in text_header_status.split() for word in list_Exclusion_Status):
            print("CANNOT BE DECONSOLIDATED - HEADER STATUS FALLS IN THE EXCLUSION LIST")
            dict_Orders = {'CASE NUMBER': [case_number], 'HPON': [order], 'COMMENTS': 'FALLOUT/HEADER STATUS IN EXCLUSION LIST'}
            df_Orders = pd.DataFrame(dict_Orders)

            session.findById("wnd[0]/tbar[0]/btn[3]").press()  # Going back to the variant "VA02"

        else:
            print("HYBRID IT ORDER")

            # Deconsolidating the order
            # Unchecking the "Complete Dlv." button
            session.findById(
                "wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/ssubHEADER_FRAME:SAPMV45A:4440/chkVBAK-AUTLF").selected = False

            session.findById("wnd[0]").sendVKey(0)  # Pressing enter

            # Check for prompt
            flag_search = True
            while flag_search:

                text_search_bar = session.findById("wnd[0]/sbar").text

                if "Delivery" in text_search_bar:
                    session.findById("wnd[0]").sendVKey(0)  # Pressing enter

                else:
                    flag_search = False

            # Going to Order Coordinator log
            session.findById("wnd[0]/mbar/menu[2]/menu[1]/menu[10]").select()

            # Adding comment in Order Coordinator Log
            existingText = session.findById(
                "wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4152/subSUBSCREEN_TEXT:SAPLV70T:2100/cntlSPLITTER_CONTAINER/shellcont/shellcont/shell/shellcont[1]/shell").text

            appendedText = "Order Deconsolidated as per case ID - " + str(case_number)

            if existingText == '\r':
                session.findById(
                    "wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4152/subSUBSCREEN_TEXT:SAPLV70T:2100/cntlSPLITTER_CONTAINER/shellcont/shellcont/shell/shellcont[1]/shell").text = appendedText
                session.findById(
                    "wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4152/subSUBSCREEN_TEXT:SAPLV70T:2100/cntlSPLITTER_CONTAINER/shellcont/shellcont/shell/shellcont[1]/shell").setSelectionIndexes(
                    8, 8)

            else:
                session.findById(
                    "wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4152/subSUBSCREEN_TEXT:SAPLV70T:2100/cntlSPLITTER_CONTAINER/shellcont/shellcont/shell/shellcont[1]/shell").text = existingText + "\n\n" + appendedText
                session.findById(
                    "wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4152/subSUBSCREEN_TEXT:SAPLV70T:2100/cntlSPLITTER_CONTAINER/shellcont/shellcont/shell/shellcont[1]/shell").setSelectionIndexes(
                    8, 8)

            # Saving the order in S4
            session.findById("wnd[0]").sendVKey(11)

            # Saving the order if it has passed all the checks
            dict_Orders = {'CASE NUMBER': [case_number], 'HPON': [order], 'COMMENTS': 'DECONSOLIDATED'}
            df_Orders = pd.DataFrame(dict_Orders)

    return df_Fallout, df_Orders


# Moving previous day file to "Archive" folder
source_folder = "Updated_Files"
destination_folder = "Updated_Files\\Archive"

files = os.listdir(source_folder)

date_pattern = r"\d{2}\.\d{2}\.\d{4}"

# dictionary to store the counts for each date
date_counts = {}

for filename in files:
    if filename.endswith(".xlsx") and re.search(date_pattern, filename):
        # Extracting the date from the filename
        date_match = re.search(date_pattern, filename)
        date = date_match.group(0)

        # Incrementing the count for the date in the dictionary
        if date in date_counts:
            date_counts[date] += 1
        else:
            date_counts[date] = 1

        # Generating the new filename with count
        count = date_counts[date]
        base_name, ext = os.path.splitext(filename)
        new_filename = f"{base_name}_{count}{ext}"

        # Checking if the new filename already exists in the destination folder
        while os.path.exists(os.path.join(destination_folder, new_filename)):
            count += 1
            new_filename = f"{base_name}_{count}{ext}"

        source_path = os.path.join(source_folder, filename)
        destination_path = os.path.join(destination_folder, new_filename)

        # Move the file to the destination folder
        shutil.move(source_path, destination_path)
        print(f"Moved {filename} to {destination_path}")


list_Exclusion_Status = ['DLRY', 'PGI', 'POD', 'INV', 'CANC']

# Creating empty DataFrames to store data from each iteration
df_Fallout_combined = pd.DataFrame()
df_Orders_combined = pd.DataFrame()

df = pd.read_excel('Raw_Files\\Raw.xlsx', sheet_name='Orders')
for i in range(0, len(df)):
    # cd_remove(df['CASE NUMBER'][i], df['HPON'][i])
    CN = df['CASE NUMBER'][i]
    HPON = df['HPON'][i]
    df_Fallout, df_Orders = cd_remove(CN, HPON)
    df_Fallout_combined = pd.concat([df_Fallout_combined, df_Fallout], ignore_index=True)
    df_Orders_combined = pd.concat([df_Orders_combined, df_Orders], ignore_index=True)

# Closing SAP Window

# Defining the SAP Logon process name
process_name = 'saplogon.exe'

# Find the process ID (PID) of the SAP Logon process
for proc in psutil.process_iter(['pid', 'name']):
    if proc.info['name'] == process_name:
        saplogon_pid = proc.info['pid']
        break
else:
    saplogon_pid = None

# Terminating the SAP Logon process if found
if saplogon_pid:
    process = psutil.Process(saplogon_pid)
    process.terminate()
    process.wait()
    print("SAP Logon window closed.")
else:
    print("SAP Logon process not found.")


# Saving the updated records
df_All_combined = pd.concat([df_Orders_combined, df_Fallout_combined], ignore_index=False)
df_All_combined.to_excel('Updated_Files\\Orders ' + var_2 + '.xlsx',
                         sheet_name='Orders', index=False)

# Making changes to the Excel file

wb = openpyxl.load_workbook('Updated_Files\\Orders ' + var_2 + '.xlsx')
ws = wb['Orders']

# Adjusting the width of the column
for column in df_All_combined:
    column_length = max(df_All_combined[column].astype(str).map(len).max(), len(column)) + 8
    col_letter = openpyxl.utils.get_column_letter(df_All_combined.columns.get_loc(column) + 1)
    ws.column_dimensions[col_letter].width = column_length

# Center aligning all cells in each column
for col in ws.columns:
    for cell in col:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

# Changing font for all cells
font = openpyxl.styles.Font(name='Calibri Light', size=11, bold=False, italic=False, color='000000')
for row in ws.iter_rows():
    for cell in row:
        cell.font = font

# Making headers bold
header_font = openpyxl.styles.Font(name='Calibri Light', size=11, bold=True, italic=False, color='000000')
for cell in ws[1]:
    cell.font = header_font

# Defining conditions and their corresponding formatting

conditions = {
    'DECONSOLIDATED': PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid'),  # Light Green
    'FALLOUT/ARUBA': PatternFill(start_color='FCE5CD', end_color='FCE5CD', fill_type='solid'),   # Light Pink
    'FALLOUT/ALLETRA|3PAR': PatternFill(start_color='FCE5CD', end_color='FCE5CD', fill_type='solid'),   # Light Pink
    'FALLOUT/HEADER STATUS IN EXCLUSION LIST': PatternFill(start_color='FCE5CD', end_color='FCE5CD', fill_type='solid'),   # Light Pink
}

# Get the column letter for the column you want to apply conditional formatting to
column_to_format = 'COMMENTS'
col_letter = openpyxl.utils.get_column_letter(df_All_combined.columns.get_loc(column_to_format) + 1)

# Loop through the cells in the column and apply conditional formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=df_All_combined.columns.get_loc(column_to_format) + 1, max_col=df_All_combined.columns.get_loc(column_to_format) + 1):
    for cell in row:
        cell_value = cell.value
        if cell_value in conditions:
            cell.fill = conditions[cell_value]

wb.save('Updated_Files\\Orders ' + var_2 + '.xlsx')
